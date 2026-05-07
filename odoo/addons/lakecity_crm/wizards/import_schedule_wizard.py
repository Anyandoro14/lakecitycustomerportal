# -*- coding: utf-8 -*-
"""Bulk import wizard for the legacy Collection Schedule XLSX templates.

Reads any of the ``Collection_Schedule_Template_*.xlsx`` files (12, 24, 36,
48, 60, 72, 84, 96, 120 months) and creates / updates the corresponding
``lakecity.collection.schedule`` records, populating monthly payment lines
from the M..FX cells.

Idempotent: a second import of the same file updates the existing rows
keyed on ``stand_number`` instead of creating duplicates.
"""
import base64
import io
import logging
import re
from datetime import date

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


CATEGORY_MAP = {
    "internal tester": "internal_tester",
    "standard": "standard",
    "vip": "vip",
    "staff": "staff",
    "partner": "partner",
    "partner / reseller": "partner",
    "reseller": "partner",
}

OPS_HEADER_MAP = {
    "present y": "present_y",
    "offer received": "offer_received",
    "initial payment completed": "initial_payment_completed",
    "agreement requested": "agreement_requested",
    "agreement signed by warwickshire": "agreement_signed_by_warwickshire",
    "agreement signed by client": "agreement_signed_by_client",
    "registered": "registered",
}

VALID_TERMS = {"12", "24", "36", "48", "60", "72", "84", "96", "120"}


def _to_bool(value):
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"true", "yes", "y", "1", "x", "✓", "done", "complete"}


def _to_date(value):
    if not value:
        return False
    if isinstance(value, date):
        return value
    try:
        return fields.Date.to_date(value)
    except (ValueError, TypeError):
        return False


class ImportScheduleWizard(models.TransientModel):
    _name = "lakecity.import.schedule.wizard"
    _description = "Import Collection Schedule from XLSX"

    file_data = fields.Binary(string="XLSX File", required=True)
    file_name = fields.Char(string="File Name")
    term_override = fields.Selection(
        [
            ("auto", "Auto-detect from sheet name / column count"),
            ("12", "12 months"),
            ("24", "24 months"),
            ("36", "36 months"),
            ("48", "48 months"),
            ("60", "60 months"),
            ("72", "72 months"),
            ("84", "84 months"),
            ("96", "96 months"),
            ("120", "120 months"),
        ],
        string="Term (override)",
        default="auto",
    )
    create_partners = fields.Boolean(
        string="Create missing customers",
        default=True,
        help="When unchecked, rows whose email/stand has no matching partner are skipped.",
    )
    log = fields.Text(string="Import Log", readonly=True)

    # ------------------------------------------------------------------
    # Action
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - environmental
            raise UserError(
                _("Python package 'openpyxl' is required to import XLSX files.")
            ) from exc

        if not self.file_data:
            raise UserError(_("Please attach an XLSX file."))

        try:
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(base64.b64decode(self.file_data)),
                data_only=True,
                read_only=False,
            )
        except Exception as exc:
            raise UserError(_("Could not parse XLSX: %s") % exc) from exc

        sheet = self._pick_data_sheet(workbook)
        if sheet is None:
            raise UserError(
                _("Could not find a 'Collection Schedule' sheet in this workbook.")
            )

        term = self._detect_term(sheet)
        log_lines = [_("Sheet: %s") % sheet.title, _("Detected term: %s months") % term]

        headers = self._read_headers(sheet)
        column_index = self._index_columns(headers)
        month_columns = self._collect_month_columns(headers)
        log_lines.append(_("Found %d monthly columns.") % len(month_columns))

        created = 0
        updated = 0
        skipped = 0

        for row_idx in range(2, sheet.max_row + 1):
            stand = sheet.cell(row_idx, column_index["stand_number"]).value
            if not stand:
                continue
            stand = str(stand).strip()

            try:
                schedule, was_created = self._upsert_schedule(
                    sheet, row_idx, column_index, term
                )
            except UserError as exc:
                log_lines.append(_("Row %d skipped: %s") % (row_idx, exc.args[0]))
                skipped += 1
                continue

            if schedule is None:
                skipped += 1
                continue

            self._apply_payment_cells(sheet, row_idx, schedule, month_columns)
            self._apply_ops_columns(sheet, row_idx, schedule, headers)

            if was_created:
                created += 1
            else:
                updated += 1

        log_lines.append(
            _("Done. Created: %d, Updated: %d, Skipped: %d.") % (created, updated, skipped)
        )
        self.log = "\n".join(log_lines)

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": self.env.context,
        }

    def action_open_schedules(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Collection Schedules"),
            "res_model": "lakecity.collection.schedule",
            "view_mode": "list,kanban,form",
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _pick_data_sheet(self, workbook):
        for name in workbook.sheetnames:
            if "collection schedule" in name.lower():
                return workbook[name]
        # Fallback: skip TEMPLATE_INSTRUCTIONS, return last data-shaped sheet.
        for name in workbook.sheetnames:
            if name.upper() != "TEMPLATE_INSTRUCTIONS":
                return workbook[name]
        return None

    def _detect_term(self, sheet):
        if self.term_override and self.term_override != "auto":
            return self.term_override
        title_match = re.search(r"(\d{1,3})\s*mo", sheet.title or "", re.IGNORECASE)
        if title_match and title_match.group(1) in VALID_TERMS:
            return title_match.group(1)
        # Fall back: count monthly columns and round to nearest valid term.
        headers = self._read_headers(sheet)
        month_count = len(self._collect_month_columns(headers))
        for term in sorted((int(t) for t in VALID_TERMS)):
            if month_count <= term:
                return str(term)
        return "120"

    def _read_headers(self, sheet):
        return {
            col: (sheet.cell(1, col).value or "")
            for col in range(1, sheet.max_column + 1)
        }

    def _index_columns(self, headers):
        idx = {}
        for col, h in headers.items():
            label = str(h).strip().lower()
            if label == "stand number":
                idx["stand_number"] = col
            elif label == "first name":
                idx["first_name"] = col
            elif label == "last name":
                idx["last_name"] = col
            elif label == "contact number":
                idx["contact_number"] = col
            elif label == "email":
                idx["email"] = col
            elif label == "customer category":
                idx["customer_category"] = col
            elif label == "documentation fee":
                idx["documentation_fee"] = col
            elif label == "deposit":
                idx["deposit"] = col
            elif label == "total price":
                idx["total_price"] = col
            elif label == "number of installments":
                idx["number_of_installments"] = col
            elif label == "payment":
                idx["payment_amount"] = col
            elif label == "start date":
                idx["start_date"] = col
            elif label == "agreement type (vat)":
                idx["agreement_type"] = col
            elif label == "receipts":
                idx["receipts_notes"] = col
        return idx

    def _collect_month_columns(self, headers):
        """Return a list of (col_index, datetime.date) for monthly headers."""
        months = []
        for col, h in headers.items():
            if isinstance(h, date):
                if h.day == 5:
                    months.append((col, h))
                continue
            if not h:
                continue
            text = str(h).strip()
            # e.g. "5 January 2022" or "5 Jan 2022"
            try:
                parsed = fields.Date.to_date(
                    self._normalise_month_header(text)
                )
                if parsed and parsed.day == 5:
                    months.append((col, parsed))
            except Exception:
                pass
        # De-duplicate / sort
        months.sort(key=lambda tup: tup[1])
        return months

    def _normalise_month_header(self, text):
        # Try several formats. fields.Date.to_date supports ISO; convert to ISO.
        from datetime import datetime as _dt

        for fmt in ("%d %B %Y", "%d %b %Y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return _dt.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return text

    def _resolve_partner(self, sheet, row_idx, column_index):
        Partner = self.env["res.partner"]
        email = self._cell(sheet, row_idx, column_index, "email")
        first = (self._cell(sheet, row_idx, column_index, "first_name") or "").strip()
        last = (self._cell(sheet, row_idx, column_index, "last_name") or "").strip()
        phone = self._cell(sheet, row_idx, column_index, "contact_number")
        full_name = (f"{first} {last}").strip() or self._cell(
            sheet, row_idx, column_index, "stand_number"
        )

        partner = Partner.browse()
        if email:
            partner = Partner.search([("email", "=ilike", email)], limit=1)
        if not partner and full_name:
            partner = Partner.search([("name", "=", full_name)], limit=1)

        if not partner:
            if not self.create_partners:
                raise UserError(
                    _("No partner found for email %s and create_partners is off.")
                    % (email or "<none>")
                )
            partner = Partner.create(
                {
                    "name": full_name,
                    "email": email or False,
                    "phone": phone or False,
                    "is_company": False,
                }
            )
        else:
            updates = {}
            if email and not partner.email:
                updates["email"] = email
            if phone and not partner.phone:
                updates["phone"] = phone
            if updates:
                partner.write(updates)
        return partner

    def _upsert_schedule(self, sheet, row_idx, column_index, term):
        Schedule = self.env["lakecity.collection.schedule"]

        stand = str(self._cell(sheet, row_idx, column_index, "stand_number") or "").strip()
        if not stand:
            return None, False

        partner = self._resolve_partner(sheet, row_idx, column_index)

        category_raw = (self._cell(sheet, row_idx, column_index, "customer_category") or "")
        category = CATEGORY_MAP.get(str(category_raw).strip().lower(), "standard")

        start_value = self._cell(sheet, row_idx, column_index, "start_date")
        start_date = _to_date(start_value)
        if not start_date:
            raise UserError(_("Missing or invalid Start Date."))

        installments = self._cell(sheet, row_idx, column_index, "number_of_installments")
        try:
            installments = int(installments) if installments else int(term)
        except (TypeError, ValueError):
            installments = int(term)

        agreement_type_raw = (
            self._cell(sheet, row_idx, column_index, "agreement_type") or ""
        )
        agreement_type = False
        atxt = str(agreement_type_raw).strip().lower()
        if "non" in atxt and "vat" in atxt:
            agreement_type = "non_vat"
        elif "vat" in atxt:
            agreement_type = "vat"

        first = (self._cell(sheet, row_idx, column_index, "first_name") or "").strip()
        last = (self._cell(sheet, row_idx, column_index, "last_name") or "").strip()

        vals = {
            "stand_number": stand,
            "partner_id": partner.id,
            "first_name": first or False,
            "last_name": last or False,
            "customer_category": category,
            "documentation_fee": self._cell(sheet, row_idx, column_index, "documentation_fee") or 0.0,
            "deposit": self._cell(sheet, row_idx, column_index, "deposit") or 0.0,
            "total_price": self._cell(sheet, row_idx, column_index, "total_price") or 0.0,
            "term_months": str(term),
            "number_of_installments": installments,
            "start_date": start_date,
            "receipts_notes": self._cell(sheet, row_idx, column_index, "receipts_notes") or False,
            "agreement_type": agreement_type,
        }

        existing = Schedule.search([("stand_number", "=", stand)], limit=1)
        if existing:
            existing.write(vals)
            return existing, False
        return Schedule.create(vals), True

    def _apply_payment_cells(self, sheet, row_idx, schedule, month_columns):
        if not month_columns:
            return
        lines_by_date = {p.due_date: p for p in schedule.payment_line_ids}
        for col, due_date in month_columns:
            value = sheet.cell(row_idx, col).value
            if value in (None, "", 0):
                continue
            try:
                amount = float(value)
            except (TypeError, ValueError):
                # e.g. a stray formula string that couldn't be evaluated
                continue
            if amount <= 0:
                continue
            line = lines_by_date.get(due_date)
            if line:
                line.write({"amount_paid": amount, "paid_date": line.paid_date or due_date})

    def _apply_ops_columns(self, sheet, row_idx, schedule, headers):
        updates = {}
        for col, header in headers.items():
            if not header:
                continue
            label = str(header).strip().lower()
            field_name = OPS_HEADER_MAP.get(label)
            if not field_name:
                continue
            updates[field_name] = _to_bool(sheet.cell(row_idx, col).value)
        if updates:
            schedule.write(updates)

    def _cell(self, sheet, row_idx, column_index, key):
        col = column_index.get(key)
        if not col:
            return None
        return sheet.cell(row_idx, col).value
