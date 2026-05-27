# -*- coding: utf-8 -*-
"""Idempotent Lake City chart of accounts sync (create missing accounts by code)."""
import logging
import os

from odoo import _

_logger = logging.getLogger(__name__)

TYPE_MAP = {
    "Current Assets": "asset_current",
    "Bank and Cash": "asset_cash",
    "Receivable": "asset_receivable",
    "Prepayments": "asset_prepayments",
    "Fixed Assets": "asset_fixed",
    "Non-current Assets": "asset_non_current",
    "Current Liabilities": "liability_current",
    "Payable": "liability_payable",
    "Non-current Liabilities": "liability_non_current",
    "Equity": "equity",
    "Current Year Earnings": "equity_unaffected",
    "Income": "income",
    "Other Income": "income_other",
    "Cost of Revenue": "expense_direct_cost",
    "Expenses": "expense",
}

EXTRA_ACCOUNTS = [
    ("121015", "Defaulted Customer Receivables", "asset_receivable", True),
    ("212080", "Customer Refunds / Cancellation Clearing", "liability_current", False),
    ("212090", "Customer Refunds Payable", "liability_current", False),
]

BANK_CURRENCY = {
    "101410": "USD",
    "101411": "ZWL",
    "101412": "USD",
    "101413": "USD",
    "101414": "USD",
    "101417": "USD",
    "101418": "CAD",
}


def _excel_paths():
    module_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(module_dir, "..", "..", "..", "..", "Account (account.account) (2).xlsx"),
        os.path.join(module_dir, "..", "..", "..", "Account (account.account) (2).xlsx"),
    ]
    for path in candidates:
        abspath = os.path.abspath(path)
        if os.path.isfile(abspath):
            return abspath
    return None


def load_coa_rows():
    xlsx = _excel_paths()
    if not xlsx:
        return list(EXTRA_ACCOUNTS)
    try:
        from openpyxl import load_workbook
    except ImportError:
        _logger.warning("openpyxl not installed; loading supplemental Lake City accounts only.")
        return list(EXTRA_ACCOUNTS)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    rows = []
    for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        code = str(row[0]).strip()
        name = "Deferred Output VAT" if code == "251020" else (row[1] or code)
        acc_type = TYPE_MAP.get(row[2], "asset_current")
        reconcile = True if acc_type == "asset_receivable" else bool(row[3])
        rows.append((code, name, acc_type, reconcile))
    wb.close()
    existing = {r[0] for r in rows}
    for item in EXTRA_ACCOUNTS:
        if item[0] not in existing:
            rows.append(item)
    return rows


def sync_lakecity_chart_of_accounts(env):
    Account = env["account.account"].sudo()
    Currency = env["res.currency"].sudo()
    companies = env["res.company"].sudo().search([])
    created = renamed = 0
    for company in companies:
        for code, name, acc_type, reconcile in load_coa_rows():
            acc = Account.with_company(company).search(
                [("code", "=", code), *Account._check_company_domain(company)],
                limit=1,
            )
            if acc:
                if code == "251020" and acc.name != name:
                    acc.write({"name": name})
                    renamed += 1
                continue
            vals = {
                "code": code,
                "name": name[:200],
                "account_type": acc_type,
                "reconcile": reconcile,
                "company_ids": [(6, 0, company.ids)],
            }
            cur_name = BANK_CURRENCY.get(code)
            if cur_name:
                cur = Currency.search([("name", "=", cur_name)], limit=1)
                if cur:
                    vals["currency_id"] = cur.id
            Account.create(vals)
            created += 1
    if created or renamed:
        _logger.info(
            "Lakecity COA sync: created %s account(s), renamed %s.",
            created,
            renamed,
        )
    return created, renamed


def ensure_stand_sales_journal(env):
    Journal = env["account.journal"].sudo()
    for company in env["res.company"].sudo().search([]):
        if Journal.search([("company_id", "=", company.id), ("code", "=", "STND")], limit=1):
            continue
        Journal.create(
            {
                "name": _("Lake City Stand Sales"),
                "code": "STND",
                "type": "general",
                "company_id": company.id,
            }
        )
