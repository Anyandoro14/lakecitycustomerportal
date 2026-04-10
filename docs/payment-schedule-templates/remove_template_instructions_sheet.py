#!/usr/bin/env python3
"""
Remove the TEMPLATE_INSTRUCTIONS sheet from every Collection_Schedule_Template_*.xlsx
under Payment Schedules - Customer Portal/.

Instructions live in COLLECTION_SCHEDULE_TEMPLATE_INSTRUCTIONS.md (repo root relative to this folder).

  docs/payment-schedule-templates/.venv-xlsx/bin/python \\
    docs/payment-schedule-templates/remove_template_instructions_sheet.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent / "Payment Schedules - Customer Portal"
SHEET = "TEMPLATE_INSTRUCTIONS"


def main() -> None:
    if not BASE.is_dir():
        raise SystemExit(f"Missing folder: {BASE}")
    paths = sorted(BASE.glob("Collection_Schedule_Template_*.xlsx"))
    for path in paths:
        wb = load_workbook(path)
        if SHEET not in wb.sheetnames:
            wb.close()
            print(f"skip {path.name}: no {SHEET!r}")
            continue
        ws = wb[SHEET]
        wb.remove(ws)
        wb.save(path)
        wb.close()
        print(f"OK {path.name}: removed {SHEET!r}")


if __name__ == "__main__":
    main()
