#!/usr/bin/env python3
"""
Populate row 2 on each Collection_Schedule_Template_*.xlsx with a shared internal tester contract.

Run from repo root (after venv from run-fix-templates.sh):

  docs/payment-schedule-templates/.venv-xlsx/bin/python \\
    docs/payment-schedule-templates/add_internal_tester_row.py
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent / "Payment Schedules - Customer Portal"

TESTER = {
    "stand": "{{########}}",
    "first": "Alex",
    "last": "Nyandoro",
    "phone": "+17785808657",
    "email": "alex@michaeltenable.com",
    "category": "Internal Tester",
    "documentation_fee": 0,
    "deposit": 5000,
    "total_price": 120_000,
    # Flexible start (5th of month); aligns with BNPL “due on the 5th”
    "start_date": datetime(2026, 5, 5),
}


def collection_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("Collection Schedule"):
            return wb[name]
    raise ValueError(f"No Collection Schedule sheet in {wb.sheetnames}")


def term_months_from_filename(path: Path) -> int:
    m = re.search(r"Template_(\d+)mo", path.name)
    if not m:
        raise ValueError(f"Cannot parse term from {path.name}")
    return int(m.group(1))


def main() -> None:
    if not BASE.is_dir():
        raise SystemExit(f"Missing folder: {BASE}")

    paths = sorted(BASE.glob("Collection_Schedule_Template_*.xlsx"))
    if not paths:
        raise SystemExit(f"No templates in {BASE}")

    for path in paths:
        n = term_months_from_filename(path)
        wb = load_workbook(path)
        ws = collection_sheet(wb)

        ws["A2"] = TESTER["stand"]
        ws["B2"] = TESTER["first"]
        ws["C2"] = TESTER["last"]
        ws["D2"] = TESTER["phone"]
        ws["E2"] = TESTER["email"]
        ws["F2"] = TESTER["category"]
        ws["G2"] = TESTER["documentation_fee"]
        ws["H2"] = TESTER["deposit"]
        ws["I2"] = TESTER["total_price"]
        ws["J2"] = n
        # Financed amount split equally: (Total price − Deposit) / N
        ws["K2"] = "=IF(J2=0,\"\",ROUND((I2-H2)/J2,2))"
        ws["L2"] = TESTER["start_date"]

        wb.save(path)
        wb.close()
        print(f"OK {path.name}: row 2 = internal tester, N={n}, PAYMENT = K2 formula")


if __name__ == "__main__":
    main()
