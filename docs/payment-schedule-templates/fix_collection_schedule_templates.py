#!/usr/bin/env python3
"""
Fix Collection Schedule Excel templates:
  - TOTAL PAID must be SUM(monthly columns only), excluding "Next Payment Column".
  - Refresh TEMPLATE_INSTRUCTIONS text (correct SUM description and column letters).

Optional: build 60-month template from 72-month if missing.

Requires: openpyxl (use the helper script so a local venv is created automatically).

From repo root:

  bash docs/payment-schedule-templates/run-fix-templates.sh

Or manually:

  python3 -m venv docs/payment-schedule-templates/.venv-xlsx
  docs/payment-schedule-templates/.venv-xlsx/bin/pip install -r docs/payment-schedule-templates/requirements-xlsx.txt
  docs/payment-schedule-templates/.venv-xlsx/bin/python docs/payment-schedule-templates/fix_collection_schedule_templates.py
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent / "Payment Schedules - Customer Portal"


def collection_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("Collection Schedule"):
            return wb[name]
    return wb[wb.sheetnames[0]]


def find_header_columns(ws):
    next_idx = total_paid_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        t = str(v).strip().lower()
        if "next payment" in t:
            next_idx = c
        if str(v).strip().upper() == "TOTAL PAID":
            total_paid_idx = c
    return next_idx, total_paid_idx


def find_deposit_and_total_columns(ws) -> tuple[int, int]:
    """Deposit + Total price columns (BNPL: balance net of deposit before instalments)."""
    dep_idx = tot_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        t = str(v).strip().lower()
        if "deposit" in t:
            dep_idx = c
        if "total price" in t:
            tot_idx = c
    if tot_idx is None:
        tot_idx = 9  # Column I — base contract price
    if dep_idx is None:
        dep_idx = 8  # Column H — deposit
    return dep_idx, tot_idx


def fix_collection_formulas(ws) -> tuple[int, int, int]:
    """Return (months_count, last_month_col_idx, next_col_idx)."""
    next_idx, total_paid_idx = find_header_columns(ws)
    if not next_idx or not total_paid_idx:
        raise ValueError(f"Could not find Next Payment / TOTAL PAID headers in {ws.title}")

    last_mo_idx = next_idx - 1
    m_idx = 13  # Column M
    months_count = last_mo_idx - m_idx + 1

    dep_idx, tot_idx = find_deposit_and_total_columns(ws)
    dep_letter = get_column_letter(dep_idx)
    tot_letter = get_column_letter(tot_idx)

    for r in range(2, ws.max_row + 1):
        m_letter = get_column_letter(m_idx)
        end_letter = get_column_letter(last_mo_idx)
        # TOTAL PAID = deposit + sum of instalment cells (M … last month)
        ws.cell(r, total_paid_idx).value = f"={dep_letter}{r}+SUM({m_letter}{r}:{end_letter}{r})"

    tp_letter = get_column_letter(total_paid_idx)

    # Current Balance = Total price - Deposit - SUM(instalments) = I - H - SUM(M:last)
    # Payment Progress = (Deposit + SUM(M:last)) / Total price
    bal_idx = prog_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        t = str(v).strip().lower()
        if t == "current balance":
            bal_idx = c
        if "payment progress" in t:
            prog_idx = c
    if bal_idx and prog_idx:
        for r in range(2, ws.max_row + 1):
            m_letter = get_column_letter(m_idx)
            end_letter = get_column_letter(last_mo_idx)
            ws.cell(r, bal_idx).value = (
                f"={tot_letter}{r}-{dep_letter}{r}-SUM({m_letter}{r}:{end_letter}{r})"
            )
            ws.cell(r, prog_idx).value = (
                f"=IF({tot_letter}{r}=0,\"\",({dep_letter}{r}+SUM({m_letter}{r}:{end_letter}{r}))/{tot_letter}{r})"
            )

    return months_count, last_mo_idx, next_idx


def update_instructions_sheet(
    wb,
    months: int,
    last_mo_letter: str,
    next_letter: str,
    tp_letter: str,
    bal_letter: str,
    prog_letter: str,
) -> None:
    if "TEMPLATE_INSTRUCTIONS" not in wb.sheetnames:
        return
    ws = wb["TEMPLATE_INSTRUCTIONS"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value)

            if "TOTAL PAID = SUM(M:{last monthly col + Next Payment Col})" in s:
                cell.value = s.replace(
                    "TOTAL PAID = SUM(M:{last monthly col + Next Payment Col})",
                    f"TOTAL PAID: =SUM(M through {last_mo_letter} on that row) — Next Payment Column excluded",
                )
                continue

            if s.startswith("This workbook is a ") and "contract template" in s:
                cell.value = f"This workbook is a {months}-month contract template."
                continue

            if "M through " in s and "Monthly payment columns" in s:
                cell.value = f"  M through {last_mo_letter}: Monthly payment columns ({months} months)"
                continue

            if re.match(r"^\s+[A-Z]{1,3}: Next Payment Column\s*$", s):
                cell.value = f"  {next_letter}: Next Payment Column"
                continue
            if re.match(r"^\s+[A-Z]{1,3}: TOTAL PAID", s):
                cell.value = f"  {tp_letter}: TOTAL PAID (formula)"
                continue
            if re.match(r"^\s+[A-Z]{1,3}: Current Balance", s):
                cell.value = f"  {bal_letter}: Current Balance (formula)"
                continue
            if re.match(r"^\s+[A-Z]{1,3}: Payment Progress", s):
                cell.value = f"  {prog_letter}: Payment Progress (formula)"
                continue

            if "Operational tracking columns" in s and re.match(r"^\s+[A-Z]{1,3}-[A-Z]{1,3}:", s):
                cell.value = f"  After {prog_letter}: operational tracking columns (Receipts … Registered)"
                continue

            if "Columns after " in s and "Receipts" in s:
                cell.value = f"  After {prog_letter}: operational tracking columns (Receipts … Registered)"
                continue


def process_workbook(path: Path) -> None:
    wb = load_workbook(path)
    ws = collection_sheet(wb)
    months, last_mo_idx, next_idx = fix_collection_formulas(ws)

    last_mo_letter = get_column_letter(last_mo_idx)
    next_letter = get_column_letter(next_idx)
    tp_letter = get_column_letter(next_idx + 1)
    bal_letter = get_column_letter(next_idx + 2)
    prog_letter = get_column_letter(next_idx + 3)

    update_instructions_sheet(wb, months, last_mo_letter, next_letter, tp_letter, bal_letter, prog_letter)
    wb.save(path)
    wb.close()
    print(
        f"OK {path.name}: {months} months → TOTAL PAID {tp_letter} = Deposit+SUM(M:{last_mo_letter}); "
        f"Balance = Total−Deposit−SUM(M:{last_mo_letter})"
    )


def build_60_month_from_72() -> Path:
    src = BASE / "Collection_Schedule_Template_72mo.xlsx"
    dst = BASE / "Collection_Schedule_Template_60mo.xlsx"
    if not src.is_file():
        raise FileNotFoundError(src)
    shutil.copy2(src, dst)

    wb = load_workbook(dst)
    ws = collection_sheet(wb)
    # Drop monthly installments 61–72 (columns 73–84 inclusive)
    ws.delete_cols(73, 12)
    ws.title = "Collection Schedule - 60mo"
    ws["J2"] = 60

    months, last_mo_idx, next_idx = fix_collection_formulas(ws)
    if months != 60:
        wb.close()
        raise RuntimeError(f"Expected 60 months after trim, got {months}")

    last_mo_letter = get_column_letter(last_mo_idx)
    next_letter = get_column_letter(next_idx)
    tp_letter = get_column_letter(next_idx + 1)
    bal_letter = get_column_letter(next_idx + 2)
    prog_letter = get_column_letter(next_idx + 3)

    update_instructions_sheet(wb, 60, last_mo_letter, next_letter, tp_letter, bal_letter, prog_letter)
    if "TEMPLATE_INSTRUCTIONS" in wb.sheetnames:
        inst = wb["TEMPLATE_INSTRUCTIONS"]
        for row in inst.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "72-month" in cell.value:
                    cell.value = cell.value.replace("72-month", "60-month")
                if isinstance(cell.value, str) and "72 months" in cell.value:
                    cell.value = cell.value.replace("72 months", "60 months")

    wb.save(dst)
    wb.close()
    print(f"Created {dst.name} from 72mo (removed 12 trailing monthly columns).")
    return dst


def main() -> None:
    if not BASE.is_dir():
        raise SystemExit(f"Missing folder: {BASE}")

    for path in sorted(BASE.glob("Collection_Schedule_Template_*.xlsx")):
        if "60mo" in path.name:
            continue
        process_workbook(path)

    p60 = BASE / "Collection_Schedule_Template_60mo.xlsx"
    if not p60.is_file():
        build_60_month_from_72()
    else:
        process_workbook(p60)

    print("Done.")


if __name__ == "__main__":
    main()
