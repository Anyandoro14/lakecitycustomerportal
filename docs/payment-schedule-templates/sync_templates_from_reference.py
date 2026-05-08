#!/usr/bin/env python3
"""
Rebuild each Collection_Schedule_Template_{N}mo.xlsx data sheet from the reference
workbook's "Collection Schedule - 36mo" tab (same layout, formulas, and column extent).

- Replaces in-sheet formula text: Collection Schedule - 36mo -> Collection Schedule - {N}mo
- Row 2: internal tester with stand = N*1000, TOTAL PRICE (I) = stand number, deposit 5000, J=N, K formula
- Clears constants on rows 3+ (keeps formulas)

Requires: openpyxl

  python3 docs/payment-schedule-templates/sync_templates_from_reference.py \\
    "/path/to/COLLECTION SCHEDULE.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

SCRIPT_DIR = Path(__file__).resolve().parent
BASE = SCRIPT_DIR / "Payment Schedules - Customer Portal"
REF_SHEET = "Collection Schedule - 36mo"
OLD_TAB_REF = "Collection Schedule - 36mo"
TERMS = (12, 24, 36, 48, 60, 72, 84, 96, 120)
# Copy only the top of the sheet (header + formulas); full1000-row workbooks are huge/slow.
MAX_TEMPLATE_ROWS = 200

M_COL = column_index_from_string("M")
FX_COL = column_index_from_string("FX")

TESTER = {
    "first": "Alex",
    "last": "Nyandoro",
    "phone": "+17785808657",
    "email": "alex@michaeltenable.com",
    "category": "Internal Tester",
    "documentation_fee": 0,
    "deposit": 5000,
    "start_date": datetime(2026, 5, 5),
}


def stand_for_term(n: int) -> int:
    return n * 1000


def copy_sheet_grid(src_ws, dest_ws) -> None:
    """Copy cell values (including formulas as strings). Styles omitted for speed/size."""
    last_row = min(src_ws.max_row, MAX_TEMPLATE_ROWS)
    for row in src_ws.iter_rows(
        min_row=1,
        max_row=last_row,
        min_col=1,
        max_col=src_ws.max_column,
    ):
        for cell in row:
            dest_ws.cell(row=cell.row, column=cell.column).value = cell.value


def rewrite_tab_in_formulas(ws, n: int) -> None:
    new_name = f"Collection Schedule - {n}mo"
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and OLD_TAB_REF in v:
                cell.value = v.replace(OLD_TAB_REF, new_name)


def clear_data_rows_keep_formulas(ws, first_row: int = 3) -> None:
    for r in range(first_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                continue
            cell.value = None


def apply_tester_row(ws, n: int) -> None:
    stand = stand_for_term(n)
    ws.cell(2, 1).value = stand
    ws.cell(2, 2).value = TESTER["first"]
    ws.cell(2, 3).value = TESTER["last"]
    ws.cell(2, 4).value = TESTER["phone"]
    ws.cell(2, 5).value = TESTER["email"]
    ws.cell(2, 6).value = TESTER["category"]
    ws.cell(2, 7).value = TESTER["documentation_fee"]
    ws.cell(2, 8).value = TESTER["deposit"]
    ws.cell(2, 9).value = stand
    ws.cell(2, 10).value = n
    ws.cell(2, 11).value = '=IF(J2=0,"",ROUND((I2-H2)/J2,2))'
    ws.cell(2, 12).value = TESTER["start_date"]
    for c in range(M_COL, FX_COL + 1):
        ws.cell(2, c).value = None
    # Operational columns after FX: clear sample customer flags/links on row 2 (FY–GB keep formulas)
    gc = column_index_from_string("GC")
    gl = column_index_from_string("GL")
    for c in range(gc, min(gl, ws.max_column) + 1):
        cell = ws.cell(2, c)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            continue
        cell.value = None


def process_template(ref_ws, path: Path, n: int) -> None:
    wb = load_workbook(path)
    tab = f"Collection Schedule - {n}mo"
    if tab not in wb.sheetnames:
        raise SystemExit(f"{path.name}: missing sheet {tab!r}")

    old = wb[tab]
    idx = wb.sheetnames.index(tab)
    wb.remove(old)
    new_ws = wb.create_sheet(tab, idx)
    copy_sheet_grid(ref_ws, new_ws)
    rewrite_tab_in_formulas(new_ws, n)
    clear_data_rows_keep_formulas(new_ws, first_row=3)
    apply_tester_row(new_ws, n)
    wb.save(path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "reference_xlsx",
        type=Path,
        help="Path to COLLECTION SCHEDULE.xlsx (must contain tab %r)" % REF_SHEET,
    )
    args = parser.parse_args()
    ref_path = args.reference_xlsx.expanduser()
    if not ref_path.is_file():
        print(f"Reference file not found: {ref_path}", file=sys.stderr)
        sys.exit(1)

    wb_ref = load_workbook(ref_path, data_only=False)
    if REF_SHEET not in wb_ref.sheetnames:
        print(f"Reference missing sheet {REF_SHEET!r}. Found: {wb_ref.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ref_ws = wb_ref[REF_SHEET]

    if not BASE.is_dir():
        print(f"Missing template folder: {BASE}", file=sys.stderr)
        sys.exit(1)

    paths = sorted(BASE.glob("Collection_Schedule_Template_*.xlsx"))
    for path in paths:
        m = re.search(r"Template_(\d+)mo", path.name)
        if not m:
            continue
        n = int(m.group(1))
        if n not in TERMS:
            print(f"Skip (unexpected term): {path.name}")
            continue
        process_template(ref_ws, path, n)
        print(f"OK {path.name}: stand {stand_for_term(n)}, formulas from reference, tab {n}mo")

    wb_ref.close()


if __name__ == "__main__":
    main()
