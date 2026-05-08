#!/usr/bin/env python3
"""sample-stands-for-spotcheck.py

Picks 10 random stand numbers across the historical Collection Schedule XLSX
files staged for the Odoo import wizard. Use the output to spot-check the
imported records in Odoo against the spreadsheet during cutover Step 5.

Usage:
    python scripts/cutover/sample-stands-for-spotcheck.py /tmp/lake-city-snapshot/

Requires openpyxl (pip install openpyxl).
"""
from __future__ import annotations

import os
import random
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print(
        "openpyxl not installed. Run: pip install openpyxl  (or activate .venv-xlsx)",
        file=sys.stderr,
    )
    sys.exit(1)


def stands_in_workbook(path: Path) -> list[tuple[str, str]]:
    """Return [(stand_number, term)] from a single XLSX. Best-effort: reads
    column A (or whichever first column has 'Stand' in row 1) starting row 2.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return []
    stand_col = None
    for i, value in enumerate(header):
        if isinstance(value, str) and "stand" in value.lower():
            stand_col = i
            break
    if stand_col is None:
        return []
    term = path.stem.replace("Collection_Schedule_Template_", "")
    out: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if stand_col >= len(row):
            continue
        value = row[stand_col]
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "stand number":
            out.append((text, term))
    return out


def main() -> int:
    if len(sys.argv) != 2:
        print(
            f"Usage: {sys.argv[0]} <directory-of-Collection_Schedule_Template_*.xlsx>",
            file=sys.stderr,
        )
        return 2

    snapshot_dir = Path(sys.argv[1])
    if not snapshot_dir.is_dir():
        print(f"Not a directory: {snapshot_dir}", file=sys.stderr)
        return 2

    all_stands: list[tuple[str, str]] = []
    for path in sorted(snapshot_dir.glob("Collection_Schedule_Template_*.xlsx")):
        stands = stands_in_workbook(path)
        print(f"  {path.name}: {len(stands)} stands", file=sys.stderr)
        all_stands.extend(stands)

    if not all_stands:
        print("No stands found. Confirm the snapshot directory has data.", file=sys.stderr)
        return 1

    random.seed(int(os.environ.get("SPOT_CHECK_SEED", "0")) or None)
    sample = random.sample(all_stands, k=min(10, len(all_stands)))

    print()
    print("Spot-check sample (10 stands):")
    print("-" * 40)
    for stand, term in sample:
        print(f"  {stand:<20} (term: {term})")
    print()
    print("For each, open in Odoo and compare against the spreadsheet:")
    print("  - customer name, sale price, term months, start date")
    print("  - total paid, current balance, number of payment lines")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
