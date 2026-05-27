#!/usr/bin/env python3
"""Regenerate data/lakecity_chart_of_accounts.xml from the repo-root Excel export."""
import os
import sys
import xml.sax.saxutils as xu

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX = os.path.join(ROOT, "Account (account.account) (2).xlsx")
OUT = os.path.join(
    ROOT,
    "odoo",
    "addons",
    "lakecity_loan_management",
    "data",
    "lakecity_chart_of_accounts.xml",
)

TYPE_MAP = {
    "Current Assets": "asset_current",
    "Bank and Cash": "asset_cash",
    "Receivable": "asset_receivable",
    "Prepayments": "asset_prepayments",
    "Fixed Assets": "asset_fixed",
    "Non-current Assets": "asset_non_current",
    "Current Liabilities": "liability_current",
    "Payable": "liability_payable",
    "Non-current Liabilities": "liability_non_current",
    "Equity": "equity",
    "Current Year Earnings": "equity_unaffected",
    "Income": "income",
    "Other Income": "income_other",
    "Cost of Revenue": "expense_direct_cost",
    "Expenses": "expense",
}

EXTRA = [
    ("121015", "Defaulted Customer Receivables", "asset_receivable", True),
    ("212080", "Customer Refunds / Cancellation Clearing", "liability_current", False),
    ("212090", "Customer Refunds Payable", "liability_current", False),
]


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    accounts = []
    for row in wb["Sheet1"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        code = str(row[0]).strip()
        name = "Deferred Output VAT" if code == "251020" else (row[1] or code)
        acc_type = TYPE_MAP[row[2]]
        reconcile = True if acc_type == "asset_receivable" else bool(row[3])
        accounts.append((code, name, acc_type, reconcile))
    wb.close()
    codes = {a[0] for a in accounts}
    for item in EXTRA:
        if item[0] not in codes:
            accounts.append(item)
    accounts.sort(key=lambda x: x[0])

    lines = ['<?xml version="1.0" encoding="utf-8"?>', '<odoo noupdate="1">']
    for code, name, acc_type, reconcile in accounts:
        lines.append('    <record id="lakecity_coa_%s" model="account.account">' % code)
        lines.append("        <field name=\"code\">%s</field>" % xu.escape(code))
        lines.append("        <field name=\"name\">%s</field>" % xu.escape(str(name)[:200]))
        lines.append("        <field name=\"account_type\">%s</field>" % acc_type)
        if reconcile:
            lines.append("        <field name=\"reconcile\">True</field>")
        lines.append("    </record>")
    lines.append("</odoo>")
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    print("Wrote %s accounts to %s" % (len(accounts), OUT))


if __name__ == "__main__":
    main()
